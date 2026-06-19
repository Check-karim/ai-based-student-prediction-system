"""Admin report generation (Excel and PDF)."""

from datetime import datetime
from io import BytesIO

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font

from config import UNIVERSITY_NAME, UNIVERSITY_SHORT


def _timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def generate_excel_report(stats: dict, students: list, predictions: list) -> BytesIO:
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)

    ws_summary["A1"] = f"{UNIVERSITY_NAME} — Admin Report"
    ws_summary["A1"].font = title_font
    ws_summary["A2"] = f"Generated: {_timestamp()}"
    ws_summary.append([])
    ws_summary.append(["Metric", "Value"])
    ws_summary["A4"].font = header_font
    ws_summary["B4"].font = header_font
    ws_summary.append(["Registered Students", stats["student_count"]])
    ws_summary.append(["Total Predictions", stats["prediction_count"]])
    ws_summary.append(["At Risk Predictions", stats["at_risk_count"]])
    ws_summary.append(["High Performers", stats["high_performer_count"]])
    ws_summary.column_dimensions["A"].width = 28
    ws_summary.column_dimensions["B"].width = 16

    ws_students = wb.create_sheet("Students")
    student_headers = [
        "Full Name",
        "Username",
        "Email",
        "Student ID",
        "Department",
        "Year",
        "Joined",
    ]
    ws_students.append(student_headers)
    for cell in ws_students[1]:
        cell.font = header_font
    for student in students:
        ws_students.append(
            [
                student.get("full_name", ""),
                student.get("username", ""),
                student.get("email", ""),
                student.get("student_id") or "",
                student.get("department") or "",
                student.get("year_of_study") or "",
                (student.get("created_at") or "")[:10],
            ]
        )
    for col in ("A", "B", "C", "D", "E", "F", "G"):
        ws_students.column_dimensions[col].width = 18

    ws_preds = wb.create_sheet("Predictions")
    pred_headers = [
        "Student",
        "Username",
        "Result",
        "Confidence %",
        "GPA",
        "Attendance %",
        "Study Hours",
        "Assignments",
        "Date",
    ]
    ws_preds.append(pred_headers)
    for cell in ws_preds[1]:
        cell.font = header_font
    for pred in predictions:
        ws_preds.append(
            [
                pred.get("full_name") or pred.get("username", ""),
                pred.get("username", ""),
                pred.get("predicted_result", ""),
                pred.get("confidence_score", ""),
                pred.get("previous_gpa", ""),
                pred.get("attendance_rate", ""),
                pred.get("study_hours", ""),
                pred.get("assignments_completed", ""),
                (pred.get("created_at") or "")[:10],
            ]
        )
    for col in ("A", "B", "C", "D", "E", "F", "G", "H", "I"):
        ws_preds.column_dimensions[col].width = 16

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_pdf_report(stats: dict, students: list, predictions: list) -> BytesIO:
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, f"{UNIVERSITY_SHORT} - Admin Report", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, f"Generated: {_timestamp()}", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(4)

    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Summary", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    summary_rows = [
        ("Registered Students", stats["student_count"]),
        ("Total Predictions", stats["prediction_count"]),
        ("At Risk Predictions", stats["at_risk_count"]),
        ("High Performers", stats["high_performer_count"]),
    ]
    for label, value in summary_rows:
        pdf.cell(90, 6, label)
        pdf.cell(0, 6, str(value), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Registered Students", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "B", 9)
    col_widths = [40, 30, 30, 25, 30, 15, 25]
    headers = ["Name", "Username", "Student ID", "Dept", "Email", "Year", "Joined"]
    for width, header in zip(col_widths, headers):
        pdf.cell(width, 7, header, border=1)
    pdf.ln()
    pdf.set_font("Helvetica", "", 8)
    if students:
        for student in students:
            row = [
                (student.get("full_name") or "")[:22],
                (student.get("username") or "")[:16],
                (student.get("student_id") or "-")[:14],
                (student.get("department") or "-")[:14],
                (student.get("email") or "")[:18],
                str(student.get("year_of_study") or "-"),
                (student.get("created_at") or "")[:10],
            ]
            for width, value in zip(col_widths, row):
                pdf.cell(width, 6, value, border=1)
            pdf.ln()
    else:
        pdf.cell(0, 6, "No students registered.", new_x="LMARGIN", new_y="NEXT")

    pdf.add_page()
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Recent Predictions", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "B", 9)
    pred_widths = [35, 30, 28, 22, 18, 22, 25]
    pred_headers = ["Student", "Username", "Result", "Conf%", "GPA", "Attend%", "Date"]
    for width, header in zip(pred_widths, pred_headers):
        pdf.cell(width, 7, header, border=1)
    pdf.ln()
    pdf.set_font("Helvetica", "", 8)
    if predictions:
        for pred in predictions:
            row = [
                (pred.get("full_name") or pred.get("username") or "")[:20],
                (pred.get("username") or "")[:16],
                (pred.get("predicted_result") or "")[:14],
                str(pred.get("confidence_score", "")),
                str(pred.get("previous_gpa", "")),
                str(pred.get("attendance_rate", "")),
                (pred.get("created_at") or "")[:10],
            ]
            for width, value in zip(pred_widths, row):
                pdf.cell(width, 6, value, border=1)
            pdf.ln()
    else:
        pdf.cell(0, 6, "No predictions recorded.", new_x="LMARGIN", new_y="NEXT")

    buffer = BytesIO()
    pdf.output(buffer)
    buffer.seek(0)
    return buffer
